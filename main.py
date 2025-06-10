# -*- coding: utf-8 -*-
"""
Civitai图片爬虫主流程：只用核心滚动代码，抓取所有图片信息并保存到Excel
同时整合testElementTree.py中的元素关系分析，找出img和button的共同祖先路径。
新增更智能的结构识别，旨在精准定位“图片-评论”单元的共同祖先。
特别增加了对“点赞数、爱心数、笑哭数、伤心数、打赏数”这5个同层级按钮的精准定位和数据提取。
根据用户提供的HTML文件（all.html, box.html, 5buttons.html）更新了选择器。
"""

import os
import json
import asyncio
import aiohttp
from datetime import datetime
import traceback
import logging
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import hashlib
import aiofiles
import re
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
import playwright._impl._errors
import time

# --- 配置 ---
PROXY = "http://127.0.0.1:10808"
TARGET_URL = "https://civitai.com/images?tags=4" # This will be overridden by KEYWORD_TARGET_FILE if used
LOG_DIR = "logs"
RESULTS_DIR = "results_civitai"
IMAGE_DIR_BASE = "images_civitai"
KEYWORD_TARGET_FILE = "urlTarget.txt"
DOWNLOAD_HISTORY_FILE = "download_history_civitai.json"

timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
log_filename = os.path.join(LOG_DIR, f"civitai_scraper_log_{timestamp}.txt")
excel_filename = os.path.join(RESULTS_DIR, f"civitai_image_results_{timestamp}.xlsx")
# 定义用于元素路径分析的日志和Excel文件
ELEMENT_LOG_FILE = os.path.join(LOG_DIR, f"element_analysis_log_{timestamp}.txt")
ELEMENT_XLSX_FILE = os.path.join(RESULTS_DIR, f"element_common_ancestors_{timestamp}.xlsx")


# --- 日志配置 ---
logger = logging.getLogger('civitai_scraper')
logger.setLevel(logging.INFO)
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)
if not os.path.exists(RESULTS_DIR):
    os.makedirs(RESULTS_DIR)
if not os.path.exists(IMAGE_DIR_BASE):
    os.makedirs(IMAGE_DIR_BASE)
file_handler = logging.FileHandler(log_filename, encoding='utf-8')
console_handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# --- 工具函数 ---
def calculate_md5(data_bytes):
    return hashlib.md5(data_bytes).hexdigest()

def load_download_history(filepath):
    global download_history
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                download_history = json.load(f)
            logger.info(f"Loaded download history from {filepath}")
        except Exception as e:
            logger.warning(f"Failed to load download history: {e}")
            download_history = {}
    else:
        logger.info(f"Download history file '{filepath}' not found. Starting with empty history.")
        download_history = {}

def save_download_history(filepath):
    global download_history
    try:
        dir_name = os.path.dirname(filepath)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(download_history, f, indent=4)
        logger.info(f"Saved download history to {filepath}")
    except Exception as e:
        logger.error(f"Error saving download history to '{filepath}': {e}\n{traceback.format_exc()}")

def read_urls_from_file(filepath):
    urls = []
    if not os.path.exists(filepath):
        logger.error(f"Error: Target URL file '{filepath}' not found. 请创建并添加URL。")
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                url = line.strip()
                if url and url.startswith("http"):
                    urls.append(url)
        if not urls:
            logger.warning(f"Warning: Target URL file '{filepath}' is empty or contains no valid URLs.")
        return urls
    except Exception as e:
        logger.error(f"Error reading URLs from '{filepath}': {e}\n{traceback.format_exc()}")
        return []

async def process_image_data(image_url, base_folder_path):
    local_filename = None
    image_content_md5 = None
    image_bytes = None
    if not image_url:
        logger.warning("Empty image URL skipped.")
        return None, None
    if image_url.startswith('http'):
        url_without_query = image_url.split('?')[0]
        file_extension = url_without_query.split('.')[-1].lower()
        if not file_extension or len(file_extension) > 5 or not file_extension.isalpha():
            file_extension = 'jpg'
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(image_url, proxy=PROXY if PROXY else None, timeout=30.0) as response:
                    response.raise_for_status()
                    image_bytes = await response.read()
                    image_content_md5 = calculate_md5(image_bytes)
                    if image_content_md5 in download_history:
                        existing_path = download_history[image_content_md5]
                        logger.info(f"Downloaded image content (MD5: {image_content_md5}) already exists at: {existing_path}. Skipping save and using existing path.")
                        return existing_path, image_content_md5
                    local_filename = os.path.join(base_folder_path, f"{image_content_md5}.{file_extension}")
                    os.makedirs(os.path.dirname(local_filename), exist_ok=True)
                    async with aiofiles.open(local_filename, 'wb') as f:
                        await f.write(image_bytes)
                    logger.info(f"Image downloaded and saved: {local_filename}")
                    download_history[image_content_md5] = local_filename
                    return local_filename, image_content_md5
        except Exception as e:
            logger.error(f"Error downloading image {image_url}: {e}")
    else:
        logger.warning(f"Unsupported image URL format (not http/https): {image_url[:100]}...")
    return None, None

# --- 从 testElementTree.py 复制并修改的函数 ---
def get_element_path(element, ancestor):
    """
    获取元素相对于指定祖先元素的路径。
    例如：div > span(2) > img
    """
    path = []
    current = element
    # 确保current和ancestor都是有效的BeautifulSoup Tag对象
    if not current or not ancestor or not hasattr(current, 'name'):
        return ""

    while current and current != ancestor:
        tag = current.name
        if not tag: # 处理没有tag的NavigableString等
            break
        # 避免无限循环，如果 current 已经没有父级但仍未达到 ancestor
        if not current.parent:
            break

        # 计算同级同名元素的索引
        siblings = [sib for sib in current.parent.find_all(tag, recursive=False)]
        if len(siblings) > 1:
            try:
                idx = siblings.index(current) + 1
                tag = f"{tag}({idx})"
            except ValueError: # 如果 current 不在 siblings 中 (不应该发生但为了健壮性)
                pass
        path.insert(0, tag)
        current = current.parent
    return " > ".join(path)

def get_common_prefix(paths):
    """
    找所有路径的最长公共前缀（以' > '为分隔）。
    paths 应该是一个列表，每个元素是 [path_string]。
    """
    if not paths:
        return ""
    # 将路径字符串拆分为列表
    split_paths = [p[0].split(" > ") for p in paths if p and p[0]]
    if not split_paths:
        return ""

    min_len = min(len(p) for p in split_paths)
    prefix = []
    for i in range(min_len):
        # 检查当前层级的所有路径是否相同
        tokens = set(p[i] for p in split_paths)
        if len(tokens) == 1:
            prefix.append(tokens.pop())
        else:
            # 如果不同，则公共前缀到此为止
            break
    return " > ".join(prefix)

def get_elements_and_paths(ancestor, tag_name):
    """
    通用函数，用于获取指定标签名下所有元素的路径（相对ancestor）。
    返回一个列表，其中每个元素是 [path_string]。
    """
    paths = []
    # 查找ancestor下的所有tag_name元素
    for elem in ancestor.find_all(tag_name):
        paths.append([get_element_path(elem, ancestor)])
    return paths

# --- 核心爬虫流程 ---
all_search_results_data = []
data_lock = asyncio.Lock()
download_history = {}
# 存储页面HTML内容，以便后续进行元素分析
global_page_html = None


async def performCivitaiImageScrape(context, target_url):
    async_name = asyncio.current_task().get_name()
    # --- Cookie 注入 ---
    try:
        if os.path.exists("cookies.json"):
            with open("cookies.json", "r", encoding="utf-8") as f:
                cookies = json.load(f)
                for cookie in cookies:
                    cookie_same_site = {'strict': 'Strict', 'Lax': 'Lax', 'none': 'None'}.get(cookie.get('sameSite'), None)
                    if cookie_same_site in ['Strict', 'Lax', 'None']:
                        cookie['sameSite'] = cookie_same_site
                    else:
                        if 'sameSite' in cookie:
                            del cookie['sameSite']
                await context.add_cookies(cookies)
            logger.info(f"{async_name} -> Cookies loaded and added to context.")
        else:
            logger.warning(f"{async_name} -> Warning: cookies.json not found. Proceeding without cookies.")
    except Exception as e:
        logger.error(f"{async_name} -> Error loading cookies: {e}")

    page = await context.new_page()
    try:
        logger.info(f"{async_name} -> Navigating to {target_url}")
        await page.goto(target_url, timeout=60000, wait_until="domcontentloaded")
        logger.info(f"{async_name} -> Successfully navigated to {target_url}")
    except Exception as e:
        logger.error(f"{async_name} -> Error navigating to {target_url}: {e}")
        await page.close()
        return

    civitai_image_folder_path = os.path.join(IMAGE_DIR_BASE, "downloaded_images")
    if not os.path.exists(civitai_image_folder_path):
        os.makedirs(civitai_image_folder_path)
        logger.info(f"{async_name} -> Created base image folder for Civitai: {civitai_image_folder_path}")

    scrape_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    processed_image_detail_urls = set()  # 放在 while 循环外
    scroll_attempts = 0
    #！！！！！
    max_scroll_attempts = 50  # 调试是50，之后再加大滚动次数
    no_new_images_start_time = None  # 新增：无新图片开始的时间

    # 根据all.html文件更新：大的容器 'all'
    main_content_area_selector = 'div.mx-auto.flex.justify-center.gap-4'

    keyword_input_selector = 'header input'  # 关键词输入框选择器
    current_keyword = "N/A"
    try:
        keyword_input_element = page.locator(keyword_input_selector)
        if await keyword_input_element.is_visible():
            input_value = await keyword_input_element.get_attribute('value')
            if input_value:
                current_keyword = input_value
                logger.info(f"{async_name} -> Found keyword in input field: '{current_keyword}'")
    except Exception as e:
        logger.warning(f"{async_name} -> Could not find or extract keyword: {e}")

    while scroll_attempts < max_scroll_attempts:
        scroll_attempts += 1
        logger.info(f"{async_name} -> Scroll attempt {scroll_attempts}/{max_scroll_attempts}...")
        # 优化滚动，尝试滚动到页面底部
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(0.05) # 短暂等待确保滚动完成

        # 在每次滚动后获取最新HTML
        page_html = await page.content()
        soup = BeautifulSoup(page_html, "html.parser")

        # 找到包含所有图片盒子的主要内容区域 'all'
        target_container = soup.select_one(main_content_area_selector)
        if not target_container:
            logger.warning(f"主内容容器未找到，选择器: {main_content_area_selector}。跳过本次循环。")
            continue

        # 在主内容容器内，识别每个独立的“图片-评论”单元 'box'
        # 根据box.html文件更新：每个图片卡片 'box'
        image_comment_units = target_container.find_all(
            "div", class_="relative flex overflow-hidden rounded-md border-gray-3 bg-gray-0 shadow-gray-4 dark:border-dark-4 dark:bg-dark-6 dark:shadow-dark-8 flex-col border"
        )
        
        current_image_count = len(image_comment_units) # 统计识别到的图片评论单元数量
        newly_processed_this_scroll = 0

        for unit in image_comment_units:
            img = unit.find("img") # img在box内部
            if not img:
                continue

            thumbnail_url = img.get("src")
            if not thumbnail_url or not thumbnail_url.startswith("http"):
                continue

            parent_a = img.find_parent("a")
            original_page_url = parent_a.get("href") if parent_a else ""
            if original_page_url and not original_page_url.startswith("http"):
                original_page_url = f"https://civitai.com{original_page_url}"

            # 去重
            unique_key = thumbnail_url + "|" + original_page_url
            if unique_key in processed_image_detail_urls:
                continue
            processed_image_detail_urls.add(unique_key)
            newly_processed_this_scroll += 1

            # --- 开始提取5个按钮的数据 ---
            likes_count = "N/A"
            love_count = "N/A"
            laugh_count = "N/A"
            sad_count = "N/A"
            tipped_count = "N/A"

            # 根据5buttons.html文件更新：5个按钮的共同父级元素
            buttons_parent = unit.select_one("div.flex.items-center.justify-center.gap-1.justify-between.p-2")

            if buttons_parent:
                buttons = buttons_parent.find_all("button")
                for btn in buttons:
                    # 提取数字，并处理可能存在的K、M等单位
                    def parse_count_text(text):
                        if not text:
                            return "0"
                        text = text.strip().replace(',', '')
                        if 'K' in text:
                            return str(int(float(text.replace('K', '')) * 1000))
                        elif 'M' in text:
                            return str(int(float(text.replace('M', '')) * 1000000))
                        try:
                            # 确保返回的是字符串形式的数字
                            return str(int(text))
                        except ValueError:
                            return "0"

                    # 尝试从mantine-Button-label内部的文本提取
                    count_label_span = btn.find("span", class_="mantine-qo1k2 flex gap-1 mantine-Button-label")
                    count_value = "0"
                    if count_label_span:
                        # 找到数字所在的文本节点，数字通常不在独立的span中，而是在父span的直接文本中
                        # 或者在 mantine-Text-root mantine-9yukw3 旁边的文本
                        # 针对 5buttons.html 发现数字在 `mantine-Button-label` 这个 span 之后，或者直接作为其文本
                        # 更好的办法是寻找数字旁边的文本节点
                        
                        # 尝试提取直接文本，例如 "👍 12111"
                        full_text = count_label_span.get_text(strip=True)
                        match = re.search(r'(\d+(\.\d+)?[KM]?)$', full_text) # 匹配末尾的数字或带KM的数字
                        if match:
                            count_value = parse_count_text(match.group(1))
                        elif count_label_span.parent and count_label_span.parent.get_text(strip=True):
                            # 有时数字在 mantine-Button-inner 的子文本中，而不是 label 内部
                            parent_text = count_label_span.parent.get_text(strip=True)
                            match = re.search(r'(\d+(\.\d+)?[KM]?)$', parent_text)
                            if match:
                                count_value = parse_count_text(match.group(1))
                                
                    # 特殊处理打赏按钮，它的数字在 mantine-Badge-label 中
                    badge_label_span = btn.find("span", class_="mantine-h9iq4m flex gap-0.5 items-center mantine-Badge-inner")
                    if badge_label_span:
                        # 找到svg后的文本，或者直接在badge_label_span中提取
                        # 这里需要更精确地找到数字，例如从span的子文本中
                        badge_text = badge_label_span.get_text(strip=True)
                        match = re.search(r'(\d+(\.\d+)?[KM]?)$', badge_text)
                        if match:
                            count_value = parse_count_text(match.group(1))
                        
                    # 根据图标或 aria-label/data-tooltip 识别按钮类型
                    # 5buttons.html 中没有 data-tooltip，但有 emoji
                    icon_div = btn.find("div", class_="mantine-Text-root mantine-9yukw3")
                    if icon_div:
                        icon_text = icon_div.get_text(strip=True)
                        if icon_text == '👍':
                            likes_count = count_value
                        elif icon_text == '❤️':
                            love_count = count_value
                        elif icon_text == '😂':
                            laugh_count = count_value
                        elif icon_text == '😢':
                            sad_count = count_value
                    # 打赏按钮没有直接的emoji，通常是svg图标
                    # 可以根据其 class 或 aria-label 来判断，或者它是第五个按钮
                    # 在 5buttons.html 中，打赏按钮的类名是 mantine-1qn9423
                    if "mantine-1qn9423" in btn.get("class", []):
                        tipped_count = count_value


            # --- 提取结束 ---

            # 下载图片（如需加速可注释掉，后续批量下载）
            local_image_path, image_md5 = await process_image_data(thumbnail_url, civitai_image_folder_path)
            if local_image_path:
                abs_path = os.path.abspath(local_image_path)
                if os.name == 'nt':
                    abs_path = abs_path.replace('\\', '/')
                    local_image_hyperlink = f"file:///{abs_path}"
                else:
                    local_image_hyperlink = f"file://{abs_path}"
            else:
                local_image_hyperlink = ""
            result_data = {
                "抓取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "搜索URL": target_url,
                "缩略图URL": thumbnail_url,
                "本地缩略图路径": os.path.abspath(local_image_path) if local_image_path else "",
                "本地缩略图超链接": local_image_hyperlink,
                "原始图片详情页链接": original_page_url,
                "点赞数": likes_count, # 添加点赞数
                "爱心数": love_count,   # 添加爱心数
                "笑哭数": laugh_count, # 添加笑哭数
                "伤心数": sad_count,   # 添加伤心数
                "打赏数": tipped_count, # 添加打赏数
                "关键词": current_keyword
            }
            async with data_lock:
                all_search_results_data.append(result_data)

        if newly_processed_this_scroll == 0 and current_image_count > 0:
            if no_new_images_start_time is None:
                no_new_images_start_time = time.time()
            elapsed = time.time() - no_new_images_start_time
            logger.info(f"{async_name} -> No new images processed this scroll. Consecutive no new images: {no_new_images_count}, elapsed: {elapsed:.1f}s")
            # 持续20秒没有新图片则停止滚动
            if elapsed >= 20:
                logger.info(f"{async_name} -> No new images for 20 seconds. Stopping scrolling.")
                break
        else:
            no_new_images_start_time = None  # 有新图片就重置
            no_new_images_count = 0 # 重置计数
    await page.close()
    logger.info(f"{async_name} -> Page closed for {target_url}.")

    # 在所有 scraping 任务完成后，再进行 HTML 页面内容的全局保存和元素分析
    # 这里假设我们只分析最后一个成功抓取到的页面的HTML
    global global_page_html
    global_page_html = page_html # 保存最后一次获取的页面HTML


# --- 元素结构分析函数 ---
async def analyze_civitai_element_structure(page_html):
    element_analysis_errors = []
    element_analysis_results = {} # 存储分析结果

    if not page_html:
        logger.warning("No page HTML content provided for element analysis. Skipping analysis.")
        return element_analysis_results, element_analysis_errors

    try:
        soup = BeautifulSoup(page_html, "html.parser")
        
        # 识别主内容区域的祖先选择器 'all'
        main_content_area_selector = 'div.mx-auto.flex.justify-center.gap-4'
        ancestor_for_elements = soup.select_one(main_content_area_selector)

        if not ancestor_for_elements:
            element_analysis_errors.append(f"Error: Ancestor element for analysis not found with selector: {main_content_area_selector}")
        else:
            # 识别每个独立的“图片-评论”单元 'box'
            image_comment_units = ancestor_for_elements.find_all(
                "div", class_="relative flex overflow-hidden rounded-md border-gray-3 bg-gray-0 shadow-gray-4 dark:border-dark-4 dark:bg-dark-6 dark:shadow-dark-8 flex-col border"
            )

            if not image_comment_units:
                element_analysis_errors.append(f"Warning: No 'image-comment' units found with selector 'div.relative.flex.overflow-hidden...' under the main ancestor.")
                
            unit_paths = []
            all_img_paths_in_units = []
            all_button_group_paths_in_units = [] # 记录按钮组的路径

            for unit in image_comment_units:
                unit_path = get_element_path(unit, ancestor_for_elements) # box相对于all
                unit_paths.append([unit_path])

                # 在每个 unit 内部查找 img
                img_in_unit = unit.find("img")
                if img_in_unit:
                    all_img_paths_in_units.append([get_element_path(img_in_unit, unit)]) # img相对于box

                # 在每个 unit 内部查找5个按钮的共同父级元素
                buttons_parent = unit.select_one("div.flex.items-center.justify-center.gap-1.justify-between.p-2") # 5button parent相对于box
                if buttons_parent:
                    all_button_group_paths_in_units.append([get_element_path(buttons_parent, unit)]) # 按钮组相对于box


            # 计算各个层级的共同祖先路径
            common_ancestor_for_all_units = get_common_prefix(unit_paths) # box的共同祖先路径 (相对于all)
            common_relative_path_for_imgs_in_units = get_common_prefix(all_img_paths_in_units) # img在box内的共同路径
            common_relative_path_for_button_groups_in_units = get_common_prefix(all_button_group_paths_in_units) # 按钮组在box内的共同路径

            element_analysis_results = {
                "Common Ancestor for Image-Comment Units (Box in All)": common_ancestor_for_all_units,
                "Common Relative Path for Images within Units (Img in Box)": common_relative_path_for_imgs_in_units,
                "Common Relative Path for Button Groups within Units (5ButtonParent in Box)": common_relative_path_for_button_groups_in_units,
                "All Image-Comment Unit Paths (Box Relative to All)": [p[0] for p in unit_paths],
                "All Relative Image Paths within Units (Img Relative to Box)": [p[0] for p in all_img_paths_in_units],
                "All Relative Button Group Paths within Units (5ButtonParent Relative to Box)": [p[0] for p in all_button_group_paths_in_units]
            }

    except Exception as e:
        element_analysis_errors.append(f"An unexpected error occurred during element analysis: {e}\n{traceback.format_exc()}")
        logger.error(f"Error during element analysis: {e}")
        
    return element_analysis_results, element_analysis_errors


# --- 主入口 ---
async def main():
    load_download_history(DOWNLOAD_HISTORY_FILE)
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            proxy={"server": PROXY} if PROXY else None,
            timeout=60000
        )
        context = await browser.new_context(
            viewport={'width': 2560, 'height': 1440}
        )
        target_urls = read_urls_from_file(KEYWORD_TARGET_FILE)
        if not target_urls:
            logger.error(f"No valid URLs found in {KEYWORD_TARGET_FILE}. Please add URLs to scrape.")
            await browser.close()
            return
        # 直接抓取所有URL
        tasks = [performCivitaiImageScrape(context, url) for url in target_urls]
        await asyncio.gather(*tasks) # 执行所有爬虫任务
        await browser.close()
        logger.info("Browser closed. Script finished scraping data.")

    # --- 元素结构分析部分 (在所有爬虫任务结束后执行一次) ---
    element_analysis_results, element_analysis_errors = await analyze_civitai_element_structure(global_page_html)

    # 写 log
    with open(ELEMENT_LOG_FILE, "w", encoding="utf-8") as f:
        if element_analysis_errors:
            f.write("\n\n".join(element_analysis_errors))
        else:
            f.write("Element analysis completed successfully.\n")
            for key, value in element_analysis_results.items():
                if isinstance(value, list):
                    f.write(f"\n{key}:\n")
                    for item in value:
                        f.write(f"  - {item}\n")
                else:
                    f.write(f"\n{key}: {value}\n")

    # 写 xlsx
    wb_elements = Workbook()
    ws_elements = wb_elements.active
    ws_elements.title = "Element Commonalities"
    
    ws_elements.append(["Analysis Category", "Value"])
    ws_elements.append(["Common Ancestor for Image-Comment Units (Box in All)", element_analysis_results.get("Common Ancestor for Image-Comment Units (Box in All)", "N/A")])
    ws_elements.append(["Common Relative Path for Images within Units (Img in Box)", element_analysis_results.get("Common Relative Path for Images within Units (Img in Box)", "N/A")])
    ws_elements.append(["Common Relative Path for Button Groups within Units (5ButtonParent in Box)", element_analysis_results.get("Common Relative Path for Button Groups within Units (5ButtonParent in Box)", "N/A")])

    # 详细路径列表
    ws_elements.append([]) # Blank row
    ws_elements.append(["Detailed Paths"])
    ws_elements.append(["Image-Comment Unit Paths (Box Relative to All)"])
    for path in element_analysis_results.get("All Image-Comment Unit Paths (Box Relative to All)", []):
        ws_elements.append([path])
    
    ws_elements.append([]) # Blank row
    ws_elements.append(["Relative Image Paths within Units (Img Relative to Box)"])
    for path in element_analysis_results.get("All Relative Image Paths within Units (Img Relative to Box)", []):
        ws_elements.append([path])

    ws_elements.append([]) # Blank row
    ws_elements.append(["Relative Button Group Paths within Units (5ButtonParent Relative to Box)"])
    for path in element_analysis_results.get("All Relative Button Group Paths within Units (5ButtonParent Relative to Box)", []):
        ws_elements.append([path])

    wb_elements.save(ELEMENT_XLSX_FILE)
    logger.info(f"Element common ancestors saved to Excel: {ELEMENT_XLSX_FILE}")

    # 自动打开 log 和 xlsx
    try:
        os.startfile(ELEMENT_LOG_FILE)
    except Exception:
        subprocess.Popen(['notepad.exe', ELEMENT_LOG_FILE])
    try:
        os.startfile(ELEMENT_XLSX_FILE)
    except Exception:
        subprocess.Popen(['start', ELEMENT_XLSX_FILE], shell=True)


    # --- Excel 导出 (主爬虫数据，增加新的列) ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Civitai图片结果"
    # 更新Headers，增加新的列
    headers = ["抓取时间", "搜索URL", "缩略图URL", "本地缩略图路径", "本地缩略图超链接", "原始图片详情页链接", "点赞数", "爱心数", "笑哭数", "伤心数", "打赏数", "关键词"]
    ws.append(headers)
    hyperlink_font = Font(color="0000FF", underline="single")
    for row_data in all_search_results_data:
        row = []
        for header in headers:
            if header == "本地缩略图超链接":
                row.append("点击打开缩略图")
            else:
                row.append(row_data.get(header, ""))
        ws.append(row)
        current_row_idx = ws.max_row
        search_url = row_data.get("搜索URL")
        if search_url:
            cell_search_url = ws.cell(row=current_row_idx, column=headers.index("搜索URL") + 1)
            cell_search_url.value = search_url
            cell_search_url.hyperlink = search_url
            cell_search_url.font = hyperlink_font
        thumbnail_url = row_data.get("缩略图URL")
        if thumbnail_url:
            cell_thumbnail_url = ws.cell(row=current_row_idx, column=headers.index("缩略图URL") + 1)
            cell_thumbnail_url.value = thumbnail_url
            cell_thumbnail_url.hyperlink = thumbnail_url
            cell_thumbnail_url.font = hyperlink_font
        local_image_hyperlink_url = row_data.get("本地缩略图超链接")
        if local_image_hyperlink_url:
            cell_local_image_hyperlink = ws.cell(row=current_row_idx, column=headers.index("本地缩略图超链接") + 1)
            cell_local_image_hyperlink.value = "点击打开缩略图"
            cell_local_image_hyperlink.hyperlink = Hyperlink(ref=local_image_hyperlink_url)
            cell_local_image_hyperlink.font = hyperlink_font
        original_page_link = row_data.get("原始图片详情页链接")
        if original_page_link:
            cell_original_page_link = ws.cell(row=current_row_idx, column=headers.index("原始图片详情页链接") + 1)
            cell_original_page_link.value = original_page_link
            cell_original_page_link.hyperlink = original_page_link
            cell_original_page_link.font = hyperlink_font
    for col_idx, header in enumerate(headers):
        max_length = len(header)
        column_letter = get_column_letter(col_idx + 1)
        for r_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=r_idx, column=col_idx + 1).value
            if cell_value:
                cell_len = len(str(cell_value))
                if cell_len > max_length:
                    max_length = cell_len
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > 100:
            adjusted_width = 100
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(excel_filename)
    logger.info(f"Results saved to Excel: {excel_filename}")
    save_download_history(DOWNLOAD_HISTORY_FILE)

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Script interrupted by user.")
    except Exception as e:
        logger.critical(f"An unhandled error occurred in main: {e}\n{traceback.format_exc()}")
    finally:
        # Closing log handlers and opening files logic from main.py
        for handler in logger.handlers[:]:
            try:
                handler.flush()
                handler.close()
                logger.removeHandler(handler)
            except Exception as e:
                print(f"Error closing log handler: {e}")
        try:
            if os.path.exists(log_filename):
                print(f"Attempting to open log file: {log_filename}")
                if os.name == 'nt':
                    os.startfile(log_filename)
                elif hasattr(os, "uname") and os.uname().sysname == 'Darwin':
                    subprocess.run(['open', log_filename])
                else:
                    subprocess.run(['xdg-open', log_filename])
            else:
                print(f"Log file not found: {log_filename}")
        except Exception as e:
            print(f"Error opening log file {log_filename}: {e}")
        try:
            if os.path.exists(excel_filename):
                print(f"Attempting to open Excel file: {excel_filename}")
                if os.name == 'nt':
                    os.startfile(excel_filename)
                elif hasattr(os, "uname") and os.uname().sysname == 'Darwin':
                    subprocess.run(['open', excel_filename])
                else:
                    subprocess.run(['xdg-open', excel_filename])
            else:
                print(f"Excel file not found: {excel_filename}")
        except Exception as e:
            print(f"Error opening Excel file {excel_filename}: {e}")